import numpy as np
import face_recognition as fr
import cv2 as cv
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Define path for images
path = 'student_images'
images = []
className = []
rollNumbers = {}

# Assuming you have a mapping of names to roll numbers
name_to_roll = {
    "jafir": 3,
    "Hasib": 10,
    "Hasan": 26,
    "Rafi": 18,
    "Masrafi": 25,
    "Sanzid": 26,
    "Rafshan": 27
    # Add other names and roll numbers here
}

# Read all images and class names
mylist = os.listdir(path)
for cl in mylist:
    curImg = cv.imread(f"{path}/{cl}")
    images.append(curImg)
    name = os.path.splitext(cl)[0]
    className.append(name)
    if name in name_to_roll:
        rollNumbers[name] = name_to_roll[name]

# Function to encode images
def encodingImages(images):
    encodelist = []
    for img in images:
        img = cv.cvtColor(img, cv.COLOR_BGR2RGB)
        encodings = fr.face_encodings(img)
        if encodings:
            encodelist.append(encodings[0])
    return encodelist

# Encode known images
encodelistKnown = encodingImages(images)
print("Encoding is complete")

# Excel file setup
excel_file = 'attendance.xlsx'
if not os.path.exists(excel_file):
    # Create a new Excel file if it doesn't exist
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "Name", "Date", "Time", "Roll Number", "Presence Indicator"])
    workbook.save(excel_file)

# Initialize webcam
capture = cv.VideoCapture(0)
id_counter = 1

while True:
    isTrue, img = capture.read()
    imgS = cv.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv.cvtColor(imgS, cv.COLOR_BGR2RGB)

    faceCurrloc = fr.face_locations(imgS)
    faceCurrEnc = fr.face_encodings(imgS, faceCurrloc)

    for encodeface, faceloc in zip(faceCurrEnc, faceCurrloc):
        matches = fr.compare_faces(encodelistKnown, encodeface)
        faceDis = fr.face_distance(encodelistKnown, encodeface)
        matchindex = np.argmin(faceDis)

        if matches[matchindex]:
            name = className[matchindex]
            roll_number = name_to_roll.get(name, "Unknown")
            print(f"Name: {name}, Roll Number: {roll_number}")

            # Get current date and time
            now = datetime.now()
            current_date = now.strftime("%Y-%m-%d")
            current_time = now.strftime("%H:%M:%S")

            # Write to Excel file
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            # Check if the name is already logged for the day
            logged_today = False
            for row in sheet.iter_rows(values_only=True):
                if row[1] == name and row[2] == current_date:
                    logged_today = True
                    break

            if not logged_today:
                sheet.append([id_counter, name, current_date, current_time, roll_number, 1])
                workbook.save(excel_file)
                id_counter += 1
                print(f"Logged {name} with Roll Number {roll_number} into the Excel file.")

    # Display webcam feed
    cv.imshow("Webcam", img)

    # Exit loop on pressing 'q'
    if cv.waitKey(1) & 0xFF == ord('q'):
        break

capture.release()
cv.destroyAllWindows()